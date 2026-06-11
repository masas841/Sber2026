# -*- coding: utf-8 -*-
"""Генерация красивой сметы в Excel для интерактивной инсталляции «Мать Чукотка»."""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# --- Палитра ---
NAVY = "1F3A5F"        # основной тёмно-синий
ACCENT = "2E75B6"      # акцентный синий
LIGHT = "DDEBF7"       # светло-голубой для шапки таблицы
ZEBRA = "F2F7FC"       # зебра для строк
TOTAL_BG = "1F3A5F"    # фон итоговой строки
WHITE = "FFFFFF"
GREY = "808080"

# --- Данные сметы ---
ITEMS = [
    ("Аренда экрана", "усл.", 1, 85000),
    ("Аренда сервера", "усл.", 1, 45000),
    ("Командировочные и транспортные расходы", "усл.", 1, 120000),
    ("Монтаж и настройка", "усл.", 1, 60000),
    ("Адаптация на английский язык, включая видеоконтент", "усл.", 1, 130000),
    ("Поддержка в течение мероприятия", "усл.", 1, 30000),
]

wb = Workbook()
ws = wb.active
ws.title = "Смета"
ws.sheet_view.showGridLines = False

# --- Ширина колонок ---
widths = {"A": 6, "B": 52, "C": 12, "D": 12, "E": 16, "F": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

thin = Side(style="thin", color="BFD7EE")
med = Side(style="medium", color=NAVY)


def cell(coord, value=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    return c


# --- Шапка документа ---
ws.merge_cells("A1:F1")
t = cell("A1", "СМЕТА")
t.font = Font(name="Calibri", size=26, bold=True, color=WHITE)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
t.fill = PatternFill("solid", fgColor=NAVY)
ws.row_dimensions[1].height = 46

ws.merge_cells("A2:F2")
s = cell("A2", "на монтаж и адаптацию интерактивной инсталляции «Мать Чукотка»")
s.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
s.fill = PatternFill("solid", fgColor=ACCENT)
ws.row_dimensions[2].height = 26

# --- Мета-информация ---
ws.merge_cells("A4:C4")
m1 = cell("A4", f"Дата составления: {date.today().strftime('%d.%m.%Y')}")
m1.font = Font(name="Calibri", size=10, color=GREY)
m1.alignment = Alignment(horizontal="left")

ws.merge_cells("D4:F4")
m2 = cell("D4", "Валюта: рубли РФ")
m2.font = Font(name="Calibri", size=10, color=GREY)
m2.alignment = Alignment(horizontal="right")
ws.row_dimensions[4].height = 16

# --- Заголовки таблицы ---
header_row = 6
headers = ["№", "Наименование работ и услуг", "Ед. изм.", "Кол-во", "Цена, ₽", "Сумма, ₽"]
for idx, h in enumerate(headers):
    col = get_column_letter(idx + 1)
    c = cell(f"{col}{header_row}", h)
    c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    c.fill = PatternFill("solid", fgColor=LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(top=med, bottom=med, left=thin, right=thin)
ws.row_dimensions[header_row].height = 30

# --- Строки таблицы ---
money_fmt = "#,##0 ₽"
first_data = header_row + 1
for i, (name, unit, qty, price) in enumerate(ITEMS):
    r = first_data + i
    zebra = ZEBRA if i % 2 else WHITE
    values = [i + 1, name, unit, qty, price, f"=D{r}*E{r}"]
    aligns = ["center", "left", "center", "center", "right", "right"]
    for j, (val, al) in enumerate(zip(values, aligns)):
        col = get_column_letter(j + 1)
        c = cell(f"{col}{r}", val)
        c.fill = PatternFill("solid", fgColor=zebra)
        c.alignment = Alignment(
            horizontal=al, vertical="center", wrap_text=(j == 1),
            indent=1 if j == 1 else 0,
        )
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.font = Font(name="Calibri", size=11, color="222222")
        if j >= 4:
            c.number_format = money_fmt
    ws.row_dimensions[r].height = 24

last_data = first_data + len(ITEMS) - 1

# --- Итоговая строка ---
total_row = last_data + 1
ws.merge_cells(f"A{total_row}:E{total_row}")
tl = cell(f"A{total_row}", "ИТОГО:")
tl.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
tl.alignment = Alignment(horizontal="right", vertical="center", indent=1)
tl.fill = PatternFill("solid", fgColor=TOTAL_BG)

tv = cell(f"F{total_row}", f"=SUM(F{first_data}:F{last_data})")
tv.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
tv.alignment = Alignment(horizontal="right", vertical="center")
tv.fill = PatternFill("solid", fgColor=TOTAL_BG)
tv.number_format = money_fmt
tv.border = Border(top=med, bottom=med, right=thin)
for col in "ABCDE":
    ws[f"{col}{total_row}"].border = Border(top=med, bottom=med, left=thin)
ws.row_dimensions[total_row].height = 32

# --- Примечание ---
note_row = total_row + 2
ws.merge_cells(f"A{note_row}:F{note_row}")
n = cell(
    f"A{note_row}",
    "Стоимость указана в рублях РФ. Цены действительны на дату составления сметы.",
)
n.font = Font(name="Calibri", size=9, italic=True, color=GREY)
n.alignment = Alignment(horizontal="left")

# --- Подписи ---
sign_row = note_row + 3
ws[f"B{sign_row}"] = "Исполнитель: _______________ / ________________ /"
ws[f"B{sign_row}"].font = Font(name="Calibri", size=10, color="222222")
ws[f"E{sign_row}"] = "Заказчик: _______________ / ________________ /"
ws[f"E{sign_row}"].font = Font(name="Calibri", size=10, color="222222")

# --- Печать ---
ws.print_area = f"A1:F{sign_row}"
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.5

OUT = "Смета_Мать_Чукотка.xlsx"
wb.save(OUT)
print(f"Готово: {OUT}")
