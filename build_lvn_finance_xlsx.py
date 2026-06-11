# -*- coding: utf-8 -*-
"""Сводная таблица: фестиваль «Лестница в небо» — возвраты и выплаты."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "Lestnica-v-nebo-finansy.xlsx"

# Данные
RETURNS_TOTAL = 5_122_806.10
RETURNS_GROUPS = [
    ("Уматурман", 1_704_545.40),
    ("Пикник", 1_000_000.00),
    ("Билеты", 2_418_260.70),
]

PAID = [
    ("Налоги", 85_000.00),
    ("Банковские комиссии", 4_790.93),
    ("Задолженности перед подрядчиками (SMM)", 155_895.00),
    ("Задолженность перед кассиром", 21_221_320.00),
    ("Задолженности перед подрядчиками (реклама)", 1_060_883.00),
    ("Бухгалтерия", 50_000.00),
]

DEBT_REMAIN = [
    ("Наина", 321_000.00),
    ("Катя", 535_000.00),
    ("Аня", 1_200_000.00),
]
DEBT_REMAIN_TOTAL = 2_056_000.00

EXPECTED = [
    ("Золото", 863_000.00, "до октября (часть раньше)"),
    ("Чайф", 1_200_000.00, "середина июля"),
    ("Билеты", 250_000.00, "акт сверки ~5 июня"),
    ("Группы, выкупленные Андреевым", 5_845_000.00, "ноябрь"),
]

ACCOUNT_BALANCE = 1_757_547.00

# Стили
THIN = Side(style="thin", color="B4C6E7")
MED = Side(style="medium", color="2F5496")
border_all = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
border_section = Border(left=THIN, right=THIN, top=MED, bottom=THIN)

FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_SECTION = PatternFill("solid", fgColor="2F5496")
FILL_SUBTOTAL = PatternFill("solid", fgColor="D6DCE4")
FILL_HIGHLIGHT = PatternFill("solid", fgColor="E2EFDA")
FILL_ALT = PatternFill("solid", fgColor="F2F7FB")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
FONT_SUB = Font(name="Calibri", size=10, italic=True, color="D9E2F3")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_MONEY = Font(name="Calibri", size=11)
FONT_MONEY_BOLD = Font(name="Calibri", size=11, bold=True)

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")

MONEY_FMT = '#,##0.00" ₽"'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, cells, fills=None, fonts=None, borders=None, aligns=None, num_cols=None):
    for col, val in enumerate(cells, 1):
        c = ws.cell(row=row, column=col, value=val)
        if fills and col <= len(fills) and fills[col - 1]:
            c.fill = fills[col - 1]
        if fonts and col <= len(fonts) and fonts[col - 1]:
            c.font = fonts[col - 1]
        if borders and col <= len(borders) and borders[col - 1]:
            c.border = borders[col - 1]
        if aligns and col <= len(aligns) and aligns[col - 1]:
            c.alignment = aligns[col - 1]
        if num_cols and col in num_cols and isinstance(val, (int, float)):
            c.number_format = MONEY_FMT


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Финансы"
    set_col_widths(ws, [42, 18, 28])

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1, "Фестиваль «Лестница в небо»")
    c.font, c.fill, c.alignment = FONT_TITLE, FILL_TITLE, ALIGN_C
    for col in range(1, 4):
        ws.cell(r, col).border = Border(bottom=MED)

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1, f"Возвраты затрат и выплаты задолженностей · {date.today().strftime('%d.%m.%Y')}")
    c.font, c.fill, c.alignment = FONT_SUB, FILL_TITLE, ALIGN_C

    def section(title, row):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, title)
        c.font, c.fill, c.alignment = FONT_SECTION, FILL_SECTION, ALIGN_L
        for col in range(1, 4):
            ws.cell(row, col).fill = FILL_SECTION
            ws.cell(row, col).border = border_section
        return row + 1

    def header_row(row):
        write_row(
            ws,
            row,
            ["Статья", "Сумма, ₽", "Примечание"],
            fills=[FILL_SUBTOTAL] * 3,
            fonts=[FONT_BOLD] * 3,
            borders=[border_all] * 3,
            aligns=[ALIGN_L, ALIGN_C, ALIGN_C],
        )
        return row + 1

    # --- Возвраты ---
    r = section("1. Возвраты затрат (получено)", 4)
    r = header_row(r)
    write_row(
        ws,
        r,
        ["Итого возвратов", RETURNS_TOTAL, ""],
        fills=[FILL_HIGHLIGHT, FILL_HIGHLIGHT, FILL_HIGHLIGHT],
        fonts=[FONT_BOLD, FONT_MONEY_BOLD, FONT_NORMAL],
        borders=[border_all] * 3,
        aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
        num_cols={2},
    )
    r += 1
    write_row(ws, r, ["в т.ч. от групп и билетов:", "", ""], fonts=[FONT_BOLD, FONT_NORMAL, FONT_NORMAL], borders=[border_all] * 3, aligns=[ALIGN_L, ALIGN_R, ALIGN_L])
    r += 1
    alt = False
    for name, amount in RETURNS_GROUPS:
        fill = FILL_ALT if alt else FILL_WHITE
        write_row(
            ws,
            r,
            [f"    {name}", amount, ""],
            fills=[fill, fill, fill],
            fonts=[FONT_NORMAL, FONT_MONEY, FONT_NORMAL],
            borders=[border_all] * 3,
            aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
            num_cols={2},
        )
        alt = not alt
        r += 1

    # --- Выплачено ---
    r = section("2. Выплачено", r + 1)
    r = header_row(r)
    paid_total = sum(a for _, a in PAID)
    alt = False
    for name, amount in PAID:
        fill = FILL_ALT if alt else FILL_WHITE
        note = ""
        if "кассир" in name.lower():
            note = "крупнейшая статья расходов"
        write_row(
            ws,
            r,
            [name, amount, note],
            fills=[fill, fill, fill],
            fonts=[FONT_NORMAL, FONT_MONEY, Font(name="Calibri", size=10, italic=True, color="666666")],
            borders=[border_all] * 3,
            aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
            num_cols={2},
        )
        alt = not alt
        r += 1
    write_row(
        ws,
        r,
        ["Итого выплачено", paid_total, ""],
        fills=[FILL_SUBTOTAL] * 3,
        fonts=[FONT_BOLD, FONT_MONEY_BOLD, FONT_NORMAL],
        borders=[border_all] * 3,
        aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
        num_cols={2},
    )
    r += 2

    # --- Остаток задолженностей ---
    r = section("3. Остаток по задолженностям (к выплате)", r)
    r = header_row(r)
    alt = False
    for name, amount in DEBT_REMAIN:
        fill = FILL_ALT if alt else FILL_WHITE
        write_row(
            ws,
            r,
            [name, amount, "подрядчики / персонал"],
            fills=[fill, fill, fill],
            fonts=[FONT_NORMAL, FONT_MONEY, Font(name="Calibri", size=10, color="666666")],
            borders=[border_all] * 3,
            aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
            num_cols={2},
        )
        alt = not alt
        r += 1
    write_row(
        ws,
        r,
        ["Итого к выплате", DEBT_REMAIN_TOTAL, ""],
        fills=[FILL_HIGHLIGHT] * 3,
        fonts=[FONT_BOLD, FONT_MONEY_BOLD, FONT_NORMAL],
        borders=[border_all] * 3,
        aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
        num_cols={2},
    )
    r += 2

    # --- Ожидаемые поступления ---
    r = section("4. Ожидаемые поступления", r)
    r = header_row(r)
    expected_total = sum(a for _, a, _ in EXPECTED)
    alt = False
    for name, amount, note in EXPECTED:
        fill = FILL_ALT if alt else FILL_WHITE
        write_row(
            ws,
            r,
            [name, amount, note],
            fills=[fill, fill, fill],
            fonts=[FONT_NORMAL, FONT_MONEY, Font(name="Calibri", size=10, color="666666")],
            borders=[border_all] * 3,
            aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
            num_cols={2},
        )
        alt = not alt
        r += 1
    write_row(
        ws,
        r,
        ["Итого ожидается", expected_total, "без учёта частичных авансов"],
        fills=[FILL_SUBTOTAL] * 3,
        fonts=[FONT_BOLD, FONT_MONEY_BOLD, FONT_NORMAL],
        borders=[border_all] * 3,
        aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
        num_cols={2},
    )
    r += 2

    # --- Сводка ---
    r = section("5. Сводка на текущий момент", r)
    net_after_debt = ACCOUNT_BALANCE - DEBT_REMAIN_TOTAL
    prospective = ACCOUNT_BALANCE + expected_total - DEBT_REMAIN_TOTAL

    summary = [
        ("Остаток на расчётном счёте", ACCOUNT_BALANCE, "факт"),
        ("Минус: остаток задолженностей (разд. 3)", -DEBT_REMAIN_TOTAL, "обязательства"),
        ("Условный остаток после выплат по разд. 3", net_after_debt, "счёт − задолженности"),
        ("Плюс: ожидаемые поступления (разд. 4)", expected_total, "план"),
        ("Прогноз с учётом поступлений и выплат разд. 3", prospective, "ориентир"),
        ("Возвраты затрат (разд. 1)", RETURNS_TOTAL, "уже получено"),
        ("Выплачено всего (разд. 2)", paid_total, "уже уплачено"),
    ]
    r = header_row(r)
    alt = False
    for name, amount, note in summary:
        fill = FILL_ALT if alt else FILL_WHITE
        if "Остаток на расчётном" in name:
            fill = FILL_HIGHLIGHT
        write_row(
            ws,
            r,
            [name, amount, note],
            fills=[fill, fill, fill],
            fonts=[FONT_BOLD if "счёте" in name or "Прогноз" in name else FONT_NORMAL, FONT_MONEY_BOLD if "счёте" in name else FONT_MONEY, Font(name="Calibri", size=10, color="666666")],
            borders=[border_all] * 3,
            aligns=[ALIGN_L, ALIGN_R, ALIGN_L],
            num_cols={2},
        )
        alt = not alt
        r += 1

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
